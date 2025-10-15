import pyautogui
import time
from openpyxl import load_workbook


# Координаты (примерные, подправить)
COORDS = {
    "fio_input": (300, 200),      # Поле ФИО
    "clear_btn": (1011, 666),     # Кнопка "Очистить"
    "add_btn": (809, 666),        # Кнопка "Добавить"
    "diagnosis_input": (440, 360) # Поле "Диагноз"
}

# Загрузка данных из Excel
wb = load_workbook('patients.xlsx')
ws = wb.active

# Пропускаем заголовок, начинаем со второй строки
for row in ws.iter_rows(min_row=2, values_only=True):
    fio, diagnosis = row[0], row[1]  # Подправить индексы под таблицу
    pyautogui.click(*COORDS["fio_input"])
    time.sleep(1)
    pyautogui.typewrite(str(fio))
    time.sleep(1)
    pyautogui.press('f2')
    time.sleep(1)
    pyautogui.doubleClick(*COORDS["fio_input"])
    time.sleep(1.5)
    pyautogui.click(*COORDS["clear_btn"])
    time.sleep(1)
    pyautogui.click(*COORDS["add_btn"])
    time.sleep(1)
    pyautogui.doubleClick(*COORDS["diagnosis_input"])
    time.sleep(1)
    pyautogui.typewrite(str(diagnosis))
    time.sleep(1)
    pyautogui.hotkey('shift', 'enter')
    time.sleep(2)
    pyautogui.press('f11')
    time.sleep(3)
